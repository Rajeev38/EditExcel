from django.shortcuts import render, redirect
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse
from openpyxl import load_workbook
import openpyxl
import os
import pandas as pd

from django.utils.datastructures import MultiValueDictKeyError


def index(request):
    return render(request, 'index.html')


def savefile(request):
    if request.method == 'POST':
        try:
            dir_name = "static/file/"
            test = os.listdir(dir_name)

            for item in test:
                os.remove(os.path.join(dir_name, item))
        except FileNotFoundError:
            pass
        try:
            file = request.FILES
            myfile = file['excelfile']
            fs = FileSystemStorage(location='static/file/')
            filename = fs.save(myfile.name, myfile)
            params = {'file': myfile.name}
            os.rename(r'static/file/{0}'.format(myfile.name), r'static/file/new.xlsx')
            return render(request, 'index.html', params)
        except MultiValueDictKeyError:
            params = {'error': 'No File Selected'}
            return render(request, 'index.html', params)


def create(request):
    if request.method == 'POST':
        if 'create' in request.POST:
            columns = request.POST['column']
            params = {'range': range(int(columns)), 'size': int(columns)}
            return render(request, 'add.html', params)

        if 'retrieve' in request.POST:
            df = pd.read_excel('static/file/new.xlsx')
            l1 = df.to_numpy().tolist()
            shape = df.shape
            params = {'data': l1, 'row': range(shape[0]), 'column': range(shape[1])}
            return render(request, 'retrieve.html', params)

        if 'update' in request.POST:
            columns = request.POST['column']
            params = {'range': range(int(columns)), 'size': int(columns)}
            return render(request, 'update.html',params)

        if 'delete' in request.POST:
            return render(request,'delete.html')


def add(request):
    if request.method == 'POST':
        size = request.POST['size']
        list1 = []
        for i in range(int(size)):
            data = request.POST[str(i)]
            list1.append(data)
        list2 = []
        list2.append(list1)
        df = pd.DataFrame(list2)
        writer = pd.ExcelWriter('static/file/new.xlsx', engine='openpyxl')
        writer.book = load_workbook('static/file/new.xlsx')
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        reader = pd.read_excel(r'static/file/new.xlsx')
        df.to_excel(writer, index=False, header=False, startrow=len(reader) + 1)

        writer.close()
        return redirect('index')


def update(request):
    if request.method == 'POST':
        size = request.POST['size']
        data = request.POST[str(0)]
        excel_file = load_workbook('static/file/new.xlsx')
        excel_sheet = excel_file['Sheet1']
        last_row = excel_sheet.max_row
        count = 0
        for i in range(last_row):
            count += 1
            if excel_sheet['A{}'.format(i)] == data:
                break
        col = 65
        for j in range(int(size)):
            if request.POST[str(0)] != None:
                excel_sheet["{}{}".format(chr(col),count)] = request.POST[str(j)]
                col += 1
        excel_file.save('static/file/new.xlsx')
    return redirect('index')

def delete(request):
    if request.method == 'POST':

        data = request.POST['key']
        excel_file = load_workbook('static/file/new.xlsx')
        excel_sheet = excel_file['Sheet1']
        last_row = excel_sheet.max_row
        count = 0
        for i in range(last_row):
            count += 1
            if excel_sheet['A{}'.format(i)] == data:
                break
        excel_sheet.delete_rows(idx=count)
        excel_file.save('static/file/new.xlsx')
    return redirect('index')

def download(request):
    with open('static/file/new.xlsx', 'rb') as fh:
        response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
        response['Content-Disposition'] = 'inline; filename=' + os.path.basename('static/file/new.xlsx')
        return response